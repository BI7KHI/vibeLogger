import datetime
import os
import json
from openpyxl import Workbook, load_workbook
from pypinyin import pinyin, Style

# 配置文件与路径
CONFIG_FILE = "log_config.json"
EXCEL_FILE = "Ham_Radio_Log_2026.xlsx"

def get_pinyin_abbr(text):
    """修正后的拼音缩写提取逻辑"""
    # pinyin 返回格式如 [['l'], ['g']]，需要提取每个子列表的第一个元素
    abbr_list = pinyin(text, style=Style.FIRST_LETTER)
    return "".join([item[0] for item in abbr_list]).lower()

def load_config():
    default_config = {
        "QTH": ["广州", "深圳", "龙岗", "南山", "福田", "宝安"],
        "Rig": ["UV-K5", "UV-K6", "森海克斯8800", "八重洲FT-65R"],
        "Power": ["5W", "10W", "25W", "50W", "100W"],
        "Antenna": ["原装天线", "老鹰775拉杆天线", "IOO天线"]
    }
    if not os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(default_config, f, ensure_ascii=False, indent=4)
        return default_config
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return default_config

def save_config(config):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

def smart_input(prompt, config_key, config_data, default_val=None, is_qth=False):
    options = config_data[config_key]
    print(f"\n>>> 选择/输入 {prompt}")
    for i, opt in enumerate(options, 1):
        abbr_hint = f" [{get_pinyin_abbr(opt)}]" if is_qth else ""
        print(f"  {i}. {opt}{abbr_hint}")
    
    hint = "序号/简拼/关键词/内容: " if is_qth else "序号/关键词/内容: "
    if default_val: hint = f"{hint[:-2]} (默认: {default_val}): "

    while True:
        user_val = input(hint).strip()
        
        if not user_val and default_val: return default_val
        if not user_val: return "N/A"
        
        # 1. 序号选择
        if user_val.isdigit():
            idx = int(user_val) - 1
            if 0 <= idx < len(options): return options[idx]
        
        # 2. 匹配逻辑
        matches = []
        if is_qth:
            # QTH: 拼音首字母精确匹配 或 字符串包含匹配
            matches = [opt for opt in options if user_val.lower() == get_pinyin_abbr(opt) or user_val.lower() in opt.lower()]
        else:
            # 其他: 字符串包含匹配
            matches = [opt for opt in options if user_val.lower() in opt.lower()]
        
        # 3. 处理匹配结果
        if len(matches) == 1:
            res = matches[0] # 取出字符串
            confirm = input(f"   ∟ 匹配到 【{res}】, 回车确认 / 输入新内容: ").strip()
            if not confirm: return res
            user_val = confirm
        
        elif len(matches) > 1:
            print(f"   ⚠️ 匹配到多个选项，请输入序号选择:")
            for m_idx, m_opt in enumerate(matches, 1):
                print(f"      {m_idx}. {m_opt}")
            sub_choice = input("   请选择序号 (或直接输入新内容): ").strip()
            if sub_choice.isdigit():
                s_idx = int(sub_choice) - 1
                if 0 <= s_idx < len(matches): return matches[s_idx]
            if sub_choice: user_val = sub_choice

        # 4. 自学习
        if user_val not in options:
            options.append(user_val)
            config_data[config_key] = options
            save_config(config_data)
            print(f"  ✨ 已学习新词汇: {user_val}")
        return user_val

def create_log():
    config = load_config()
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE); ws = wb.active
        except:
            print("\n❌ 错误：Excel 文件正在打开，请关闭后运行！"); return
    else:
        wb = Workbook(); ws = wb.active; ws.title = "点名日志"
        ws.append(["序号", "时间", "呼号", "QTH", "信号报告", "设备", "功率", "天馈", "留言"])

    print("="*55)
    print("   业余无线电台网日志助手 2026 (智能混合匹配版)   ")
    print("="*55)

    while True:
        next_seq = ws.max_row
        current_time = datetime.datetime.now().strftime("%H:%M")
        print(f"\n【No.{next_seq} | {current_time}】")

        call_in = input("请输入呼号 (Callsign): ").strip()
        if not call_in: continue
        callsign = call_in.upper()

        qth = smart_input("QTH (所在地)", "QTH", config, is_qth=True)
        rst = input("请输入 RST [默认 59]: ").strip() or "59"
        rig = smart_input("设备 (Rig)", "Rig", config)
        pwr = smart_input("功率 (Power)", "Power", config, default_val="5W")
        ant = smart_input("天馈 (Antenna)", "Antenna", config)
        msg = input("讨论话题及留言 [默认 73]: ").strip() or "73"

        ws.append([next_seq, current_time, callsign, qth, rst, rig, pwr, ant, msg])
        wb.save(EXCEL_FILE)
        
        # 回显核对
        print("-" * 35)
        print(f"✅ 已记录：{callsign} | {qth} | {rig} | {pwr} | {ant} | {rst} | {msg}")
        print("-" * 35)

        if input("\n[回车] 下一位，[n] 退出: ").lower() == 'n': break

if __name__ == "__main__":
    create_log()
