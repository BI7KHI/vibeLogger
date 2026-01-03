import datetime
import os
import json
import csv
import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import Workbook, load_workbook
from pypinyin import pinyin, Style

CONFIG_FILE = "log_config.json"
EXCEL_FILE = "Ham_Radio_Log_2026.xlsx"
CSV_FILE = "Ham_Radio_Log_2026.csv"


def get_pinyin_abbr(text: str) -> str:
    """拼音首字母缩写"""
    if not text:
        return ""
    abbr_list = pinyin(text, style=Style.FIRST_LETTER)
    return "".join([item[0] for item in abbr_list]).lower()


def load_config():
    default_config = {
        "QTH": ["广州", "深圳", "龙岗", "南山", "福田", "宝安"],
        "Rig": ["UV-K5", "UV-K6", "森海克斯8800", "八重洲FT-65R"],
        "Power": ["5W", "10W", "25W", "50W", "100W"],
        "Antenna": ["原装天线", "老鹰775拉杆天线", "IOO天线"],
    }
    if not os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(default_config, f, ensure_ascii=False, indent=4)
        return default_config

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default_config


def save_config(config):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)


def init_workbook():
    """打开或创建 Excel 工作簿，结构与命令行版本一致"""
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except Exception:
            messagebox.showerror("错误", "Excel 文件正在打开，请先关闭后再运行！")
            return None, None
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "点名日志"
        ws.append(["序号", "时间", "呼号", "QTH", "信号报告", "设备", "功率", "天馈", "留言"])
        wb.save(EXCEL_FILE)
    return wb, ws


def export_to_csv(ws):
    """将当前工作表完整导出为 CSV 文件（含表头）。"""
    try:
        with open(CSV_FILE, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(list(row))
    except Exception as e:
        # 导出失败只在状态栏提示，不中断主流程
        print("导出 CSV 失败:", e)


class VibeLoggerGUI:
    """业余无线电台网日志助手 GUI 版，与命令行版本字段一致"""
    
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("业余无线电台网日志助手 2026 (智能混合匹配版 · GUI)")
        self.root.geometry("1200x700")  # 增大窗口以容纳终端

        self.config = load_config()
        self.wb, self.ws = init_workbook()
        if self.wb is None:
            self.root.destroy()
            return

        self.seq_var = tk.StringVar()
        self.time_var = tk.StringVar()

        self.callsign_var = tk.StringVar()
        self.qth_var = tk.StringVar()
        self.rst_var = tk.StringVar(value="59")
        self.rig_var = tk.StringVar()
        self.power_var = tk.StringVar(value="5W")
        self.ant_var = tk.StringVar()
        self.msg_text = None

        self.status_var = tk.StringVar()
        # 日志表格控件引用
        self.log_tree = None
        
        # 命令行录入模式状态
        self.cli_log_mode = False
        self.cli_log_step = ""
        self.cli_log_data = {}
        self.current_matches = []  # 存储当前匹配的选项

        self.build_ui()
        self.refresh_header()
        self.load_existing_logs_into_view()
        # 启动时间更新
        self.update_time()

    def update_time(self):
        """实时更新时间显示，准确到秒"""
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.time_var.set(current_time)
        # 每1000毫秒（1秒）更新一次
        self.root.after(1000, self.update_time)

    def build_ui(self):
        title = tk.Label(
            self.root,
            text="业余无线电台网日志助手 2026 (智能混合匹配版 · GUI)",
            font=("微软雅黑", 14, "bold"),
        )
        title.pack(pady=6)

        header = tk.Frame(self.root)
        header.pack(fill="x", padx=20, pady=4)

        # 左侧序号显示
        left_frame = tk.Frame(header)
        left_frame.pack(side="left")
        
        tk.Label(left_frame, text="序号：", font=("微软雅黑", 10)).pack(side="left")
        tk.Label(left_frame, textvariable=self.seq_var, font=("Consolas", 11, "bold")).pack(
            side="left", padx=(0, 20)
        )

        # 右侧时间显示
        right_frame = tk.Frame(header)
        right_frame.pack(side="right")
        
        tk.Label(right_frame, text="当前时间：", font=("微软雅黑", 12)).pack(side="left", padx=(0, 5))
        tk.Label(right_frame, textvariable=self.time_var, font=("Consolas", 16, "bold"), fg="blue").pack(
            side="left"
        )

        # 主内容区域：左右分割
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill="both", expand=True, padx=20, pady=8)

        # 左侧表单区域
        form = tk.Frame(main_frame)
        form.pack(side="left", fill="both", expand=False, padx=(0, 10))

        # 右侧命令行终端区域
        terminal_frame = tk.LabelFrame(main_frame, text="命令行终端", padx=5, pady=5)
        terminal_frame.pack(side="right", fill="both", expand=True)        # 终端输出显示区域
        self.terminal_output = tk.Text(
            terminal_frame,
            height=15,
            width=50,
            bg="black",
            fg="#00FF00",  # 亮绿色，提高可读性
            font=("Consolas", 9),
            state="disabled"
        )
        self.terminal_output.pack(fill="both", expand=True, pady=(0, 5))

        # 终端滚动条
        terminal_scrollbar = ttk.Scrollbar(terminal_frame, command=self.terminal_output.yview)
        terminal_scrollbar.pack(side="right", fill="y")
        self.terminal_output.configure(yscrollcommand=terminal_scrollbar.set)

        # 命令输入框
        terminal_input_frame = tk.Frame(terminal_frame)
        terminal_input_frame.pack(fill="x", pady=(0, 5))
        
        tk.Label(terminal_input_frame, text=">", font=("Consolas", 10, "bold"), fg="#00FF00").pack(side="left")
        self.terminal_input = tk.Entry(
            terminal_input_frame,
            font=("Consolas", 9),
            bg="black",
            fg="#00FF00",  # 亮绿色，提高可读性
            insertbackground="#00FF00"  # 光标颜色也改为亮绿色
        )
        self.terminal_input.pack(side="left", fill="x", expand=True, padx=(5, 0))
        self.terminal_input.bind("<Return>", self.execute_command)
        
        # 绑定全局键盘快捷键
        self.root.bind("<Control-c>", self.handle_ctrl_c)
        # 确保终端输入框也能响应快捷键
        self.terminal_input.bind("<Control-c>", self.handle_ctrl_c)

        # 初始化终端
        self.terminal_history = []
        self.history_index = -1
        self.print_to_terminal("业余无线电台网日志助手命令行终端")
        self.print_to_terminal("输入 'help' 查看可用命令")
        self.print_to_terminal("按 Ctrl+C 可随时退出命令行状态")
        self.print_to_terminal("=" * 40)

        row = 0
        # 呼号
        tk.Label(form, text="呼号 (Callsign)：", anchor="e", width=16).grid(
            row=row, column=0, sticky="e", pady=4
        )
        tk.Entry(form, textvariable=self.callsign_var, width=22).grid(
            row=row, column=1, sticky="w", pady=4
        )
        row += 1

        # QTH：带简拼匹配
        tk.Label(form, text="QTH (所在地)：", anchor="e", width=16).grid(
            row=row, column=0, sticky="e", pady=4
        )
        self.qth_combo = ttk.Combobox(
            form,
            textvariable=self.qth_var,
            values=self.config.get("QTH", []),
            width=20,
        )
        self.qth_combo.grid(row=row, column=1, sticky="w", pady=4)
        self.qth_combo.bind("<KeyRelease>", self.on_qth_typing)
        tk.Label(form, text="可输简拼/关键字：gz、sz、lg…", fg="gray").grid(
            row=row, column=2, sticky="w"
        )
        row += 1

        # RST
        tk.Label(form, text="RST：", anchor="e", width=16).grid(
            row=row, column=0, sticky="e", pady=4
        )
        tk.Entry(form, textvariable=self.rst_var, width=10).grid(
            row=row, column=1, sticky="w", pady=4
        )
        tk.Label(form, text="默认 59", fg="gray").grid(row=row, column=2, sticky="w")
        row += 1

        # Rig
        tk.Label(form, text="设备 (Rig)：", anchor="e", width=16).grid(
            row=row, column=0, sticky="e", pady=4
        )
        self.rig_combo = ttk.Combobox(
            form,
            textvariable=self.rig_var,
            values=self.config.get("Rig", []),
            width=20,
        )
        self.rig_combo.grid(row=row, column=1, sticky="w", pady=4)
        tk.Label(form, text="支持新内容，自动学习保存", fg="gray").grid(
            row=row, column=2, sticky="w"
        )
        row += 1

        # Power
        tk.Label(form, text="功率 (Power)：", anchor="e", width=16).grid(
            row=row, column=0, sticky="e", pady=4
        )
        self.power_combo = ttk.Combobox(
            form,
            textvariable=self.power_var,
            values=self.config.get("Power", []),
            width=20,
        )
        self.power_combo.grid(row=row, column=1, sticky="w", pady=4)
        tk.Label(form, text="默认 5W，可自定义", fg="gray").grid(row=row, column=2, sticky="w")
        row += 1

        # 天馈
        tk.Label(form, text="天馈 (Antenna)：", anchor="e", width=16).grid(
            row=row, column=0, sticky="e", pady=4
        )
        self.ant_combo = ttk.Combobox(
            form,
            textvariable=self.ant_var,
            values=self.config.get("Antenna", []),
            width=20,
        )
        self.ant_combo.grid(row=row, column=1, sticky="w", pady=4)
        tk.Label(form, text="支持新内容，自动学习保存", fg="gray").grid(
            row=row, column=2, sticky="w"
        )
        row += 1

        # 留言
        tk.Label(form, text="留言 / 话题：", anchor="ne", width=16).grid(
            row=row, column=0, sticky="ne", pady=4
        )
        self.msg_text = tk.Text(form, width=48, height=4)
        self.msg_text.grid(row=row, column=1, columnspan=2, sticky="w", pady=4)
        self.msg_text.insert("1.0", "73")
        row += 1

        # 按钮
        btns = tk.Frame(self.root)
        btns.pack(pady=6)
        tk.Button(btns, text="保存当前记录", width=16, command=self.save_record).pack(
            side="left", padx=10
        )
        tk.Button(btns, text="清空 / 下一位", width=16, command=self.next_record).pack(
            side="left", padx=10
        )
        tk.Button(btns, text="退出", width=10, command=self.root.quit).pack(
            side="left", padx=10
        )

        # 日志显示区域
        log_frame = tk.LabelFrame(self.root, text="通联日志", padx=5, pady=5)
        log_frame.pack(fill="both", expand=True, padx=10, pady=(0, 5))

        columns = ("seq", "time", "callsign", "qth", "rst", "rig", "power", "ant", "msg")
        self.log_tree = ttk.Treeview(
            log_frame,
            columns=columns,
            show="headings",
            height=8,
        )
        self.log_tree.pack(side="left", fill="both", expand=True)

        self.log_tree.heading("seq", text="序号")
        self.log_tree.heading("time", text="时间")
        self.log_tree.heading("callsign", text="呼号")
        self.log_tree.heading("qth", text="QTH")
        self.log_tree.heading("rst", text="RST")
        self.log_tree.heading("rig", text="设备")
        self.log_tree.heading("power", text="功率")
        self.log_tree.heading("ant", text="天馈")
        self.log_tree.heading("msg", text="留言")

        # 设置各列宽度（可根据需要微调）
        self.log_tree.column("seq", width=50, anchor="center")
        self.log_tree.column("time", width=70, anchor="center")
        self.log_tree.column("callsign", width=80, anchor="center")
        self.log_tree.column("qth", width=80, anchor="center")
        self.log_tree.column("rst", width=50, anchor="center")
        self.log_tree.column("rig", width=110, anchor="w")
        self.log_tree.column("power", width=70, anchor="center")
        self.log_tree.column("ant", width=120, anchor="w")
        self.log_tree.column("msg", width=200, anchor="w")

        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_tree.configure(yscrollcommand=scrollbar.set)

        status = tk.Label(
            self.root,
            textvariable=self.status_var,
            bd=1,
            relief="sunken",
            anchor="w",
        )
        status.pack(side="bottom", fill="x")

    def refresh_header(self):
        if not self.ws:
            return
        next_seq = self.ws.max_row
        self.seq_var.set(str(next_seq))

    def on_qth_typing(self, event):
        text = self.qth_var.get().strip()
        base = self.config.get("QTH", [])
        if not text:
            self.qth_combo["values"] = base
            return
        lower = text.lower()
        matches = []
        for opt in base:
            abbr = get_pinyin_abbr(opt)
            if lower == abbr or lower in opt.lower():
                matches.append(opt)
        self.qth_combo["values"] = matches or base

    def learn_new_value(self, key: str, value: str):
        if not value:
            return
        items = self.config.get(key, [])
        if value not in items:
            items.append(value)
            self.config[key] = items
            save_config(self.config)
            if key == "QTH":
                self.qth_combo["values"] = items
            elif key == "Rig":
                self.rig_combo["values"] = items
            elif key == "Power":
                self.power_combo["values"] = items
            elif key == "Antenna":
                self.ant_combo["values"] = items
            self.status_var.set(f"✨ 已学习新词汇: {key} -> {value}")

    def load_existing_logs_into_view(self):
        if not self.ws or not self.log_tree:
            return
        # 跳过表头，从第二行开始
        for row in list(self.ws.iter_rows(min_row=2, values_only=True)):
            if row[0] is not None:  # 确保行不为空
                seq, t, callsign, qth, rst, rig, power, ant, msg = row
                self.log_tree.insert("", "end", values=(seq, t, callsign, qth, rst, rig, power, ant, msg))

    def save_record(self):
        if not self.ws:
            messagebox.showerror("错误", "工作表未初始化！")
            return
        callsign = self.callsign_var.get().strip().upper()
        if not callsign:
            messagebox.showwarning("提示", "呼号不能为空！")
            return

        qth = self.qth_var.get().strip()
        rst = self.rst_var.get().strip() or "59"
        rig = self.rig_var.get().strip()
        power = self.power_var.get().strip() or "5W"
        ant = self.ant_var.get().strip()
        msg = self.msg_text.get("1.0", "end").strip() or "73"

        self.learn_new_value("QTH", qth)
        self.learn_new_value("Rig", rig)
        self.learn_new_value("Power", power)
        self.learn_new_value("Antenna", ant)

        next_seq = self.ws.max_row
        current_time = datetime.datetime.now().strftime("%H:%M")

        self.ws.append([next_seq, current_time, callsign, qth, rst, rig, power, ant, msg])
        self.wb.save(EXCEL_FILE)

        # 同步导出 CSV
        export_to_csv(self.ws)

        # 在页面下方的日志表格追加一行
        if self.log_tree is not None:
            self.log_tree.insert(
                "",
                "end",
                values=(next_seq, current_time, callsign, qth, rst, rig, power, ant, msg),
            )
            # 自动滚动到最新一行
            children = self.log_tree.get_children()
            if children:
                self.log_tree.see(children[-1])

        self.status_var.set(
            f"✅ 已记录：{callsign} | {qth} | {rig} | {power} | {ant} | {rst} | {msg}"
        )
        messagebox.showinfo(
            "已保存",
            f"已记录：{callsign} | {qth} | {rig} | {power} | {ant} | {rst} | {msg}",
        )

        self.next_record(auto_from_save=True)

    def next_record(self, auto_from_save: bool = False):
        self.callsign_var.set("")
        self.rst_var.set("59")
        self.msg_text.delete("1.0", "end")
        self.msg_text.insert("1.0", "73")
        self.refresh_header()
        if not auto_from_save:
            self.status_var.set("已清空输入，可以录入下一位。")    # ===== 终端相关方法 =====

    def handle_ctrl_c(self, event):
        """处理 Ctrl+C 快捷键，退出命令行状态"""
        if self.cli_log_mode:
            # 如果正在录入模式，退出录入模式
            self.cli_log_mode = False
            self.cli_log_step = ""
            self.cli_log_data = {}
            self.current_matches = []
            self.print_to_terminal("\n^C")
            self.print_to_terminal("已退出录入模式")
            self.print_to_terminal("=" * 40)
        else:
            # 如果不在录入模式，清空当前输入
            self.terminal_input.delete(0, tk.END)
            self.print_to_terminal("^C")
        
        # 阻止事件继续传播
        return "break"

    def print_to_terminal(self, message):
        """在终端输出区域打印消息"""
        self.terminal_output.config(state="normal")
        self.terminal_output.insert(tk.END, f"{message}\n")
        self.terminal_output.config(state="disabled")
        self.terminal_output.see(tk.END)

    def execute_command(self, event):
        """执行命令行输入的命令"""
        command = self.terminal_input.get().strip()
        if not command:
            return

        # 清空输入框
        self.terminal_input.delete(0, tk.END)

        # 检查是否在录入模式
        if self.cli_log_mode:
            self.print_to_terminal(f"> {command}")
            self.process_cli_log_input(command)
            return

        # 添加到历史记录
        self.terminal_history.append(command)
        self.history_index = len(self.terminal_history)

        # 显示输入的命令
        self.print_to_terminal(f"> {command}")        # 处理命令
        self.process_command(command)

    def process_command(self, command):
        """处理具体的命令"""
        cmd_lower = command.lower()
        
        if cmd_lower == "help":
            self.print_to_terminal("可用命令:")
            self.print_to_terminal("  help        - 显示帮助信息")
            self.print_to_terminal("  status      - 显示当前状态")
            self.print_to_terminal("  count       - 显示记录总数")
            self.print_to_terminal("  list [n]    - 显示最近n条记录（默认5条）")
            self.print_to_terminal("  clear       - 清空终端")
            self.print_to_terminal("  save        - 保存当前记录")
            self.print_to_terminal("  reset       - 重置输入表单")
            self.print_to_terminal("  log         - 进入命令行录入模式")
            self.print_to_terminal("")
            self.print_to_terminal("快捷键:")
            self.print_to_terminal("  Ctrl+C      - 退出录入模式或清空输入")
            
        elif cmd_lower == "status":
            self.print_to_terminal(f"当前序号: {self.seq_var.get()}")
            self.print_to_terminal(f"时间: {self.time_var.get()}")
            self.print_to_terminal(f"已录入记录数: {max(0, int(self.seq_var.get()) - 1)}")
            
        elif cmd_lower == "count":
            if self.ws:
                count = self.ws.max_row - 1  # 减去表头
                self.print_to_terminal(f"总记录数: {count}")
            else:
                self.print_to_terminal("工作表未初始化")
                
        elif cmd_lower.startswith("list"):
            parts = cmd_lower.split()
            n = 5  # 默认显示5条
            if len(parts) > 1:
                try:
                    n = int(parts[1])
                except:
                    n = 5
            self.show_recent_records(n)
            
        elif cmd_lower == "clear":
            self.terminal_output.config(state="normal")
            self.terminal_output.delete(1.0, tk.END)
            self.terminal_output.config(state="disabled")
            self.print_to_terminal("终端已清空")
            
        elif cmd_lower == "save":
            self.save_record()
            self.print_to_terminal("已执行保存操作")
            
        elif cmd_lower == "reset":
            self.next_record()
            self.print_to_terminal("已重置输入表单")
            
        elif cmd_lower == "log":
            self.start_cli_log_mode()
            
        else:
            self.print_to_terminal(f"未知命令: {command}")
            self.print_to_terminal("输入 'help' 查看可用命令")

    def show_recent_records(self, n):
        """显示最近n条记录"""
        if not self.ws:
            self.print_to_terminal("工作表未初始化")
            return
            
        rows = list(self.ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            self.print_to_terminal("暂无记录")
            return
            
        recent_rows = rows[-n:] if len(rows) > n else rows
        self.print_to_terminal(f"最近 {len(recent_rows)} 条记录:")
        self.print_to_terminal("-" * 60)
        for row in recent_rows:
            if row[0] is not None:
                seq, time, callsign, qth, rst, rig, power, ant, msg = row
                self.print_to_terminal(f"{seq:2} | {time} | {callsign:8} | {qth:6} | {rst:2} | {rig}")

    # ===== 命令行录入模式 =====

    def start_cli_log_mode(self):
        """启动命令行录入模式"""
        if not self.ws:
            self.print_to_terminal("❌ 错误：工作表未初始化")
            return
            
        self.print_to_terminal("=" * 50)
        self.print_to_terminal("进入命令行录入模式")
        self.print_to_terminal("按步骤输入各字段信息，输入 'exit' 退出录入模式")
        self.print_to_terminal("=" * 50)
        
        # 设置录入模式状态
        self.cli_log_mode = True
        self.cli_log_step = "callsign"
        self.cli_log_data = {}
        
        next_seq = self.ws.max_row
        current_time = datetime.datetime.now().strftime("%H:%M")
        self.print_to_terminal(f"【No.{next_seq} | {current_time}】")
        self.print_to_terminal("请输入呼号 (Callsign):")

    def smart_match_input(self, user_input, config_key, is_qth=False):
        """智能匹配输入，类似原版命令行的逻辑"""
        options = self.config.get(config_key, [])
        user_val = user_input.strip()
        
        # 处理空输入
        if not user_val:
            return None
        
        # 1. 序号选择
        if user_val.isdigit():
            idx = int(user_val) - 1
            if 0 <= idx < len(options):
                result = options[idx]
                self.print_to_terminal(f"   ∟ 选择了 【{result}】")
                return result
        
        # 2. 匹配逻辑
        matches = []
        if is_qth:
            # QTH: 拼音首字母精确匹配 或 字符串包含匹配
            matches = [opt for opt in options if user_val.lower() == get_pinyin_abbr(opt) or user_val.lower() in opt.lower()]
        else:
            # 其他: 字符串包含匹配
            matches = [opt for opt in options if user_val.lower() in opt.lower()]
        
        # 3. 处理匹配结果
        if len(matches) == 1:
            result = matches[0]
            self.print_to_terminal(f"   ∟ 匹配到 【{result}】")
            return result
        elif len(matches) > 1:
            self.print_to_terminal("   ⚠️ 匹配到多个选项:")
            for m_idx, m_opt in enumerate(matches, 1):
                self.print_to_terminal(f"      {m_idx}. {m_opt}")
            self.print_to_terminal("请输入序号选择，或输入新内容")
            # 存储匹配结果供后续使用
            self.current_matches = matches
            return "MULTIPLE_MATCHES"
        
        # 4. 无匹配，使用用户输入
        return user_val

    def show_options_for_input(self, config_key, prompt_text, is_qth=False):
        """显示选项列表"""
        options = self.config.get(config_key, [])
        self.print_to_terminal(f"\n>>> {prompt_text}")
        for i, opt in enumerate(options, 1):
            abbr_hint = f" [{get_pinyin_abbr(opt)}]" if is_qth else ""
            self.print_to_terminal(f"  {i}. {opt}{abbr_hint}")
        
        hint = "序号/简拼/关键词/内容" if is_qth else "序号/关键词/内容"
        self.print_to_terminal(f"请输入 {hint}:")

    def process_cli_log_input(self, user_input):
        """处理命令行录入模式的输入"""
        if user_input.lower() == "exit":
            self.cli_log_mode = False
            self.print_to_terminal("已退出录入模式")
            return
        
        step = self.cli_log_step
        
        if step == "callsign":
            if user_input.strip():
                self.cli_log_data["callsign"] = user_input.strip().upper()
                self.cli_log_step = "qth"
                self.show_options_for_input("QTH", "选择/输入 QTH (所在地)", is_qth=True)
            else:
                self.print_to_terminal("呼号不能为空，请重新输入:")
                
        elif step == "qth":
            if not user_input.strip():  # 空输入处理
                self.print_to_terminal("QTH不能为空，请重新输入:")
                return
                
            result = self.smart_match_input(user_input, "QTH", is_qth=True)
            if result == "MULTIPLE_MATCHES":
                self.cli_log_step = "qth_select"
            elif result:
                self.cli_log_data["qth"] = result
                self.cli_log_step = "rst"
                self.print_to_terminal("请输入 RST [默认 59]:")
                
        elif step == "qth_select":
            # 处理多选情况
            if user_input.isdigit():
                idx = int(user_input) - 1
                if 0 <= idx < len(self.current_matches):
                    self.cli_log_data["qth"] = self.current_matches[idx]
                    self.cli_log_step = "rst"
                    self.print_to_terminal("请输入 RST [默认 59]:")
                else:
                    self.print_to_terminal("序号无效，请重新选择:")
            else:
                # 直接使用用户输入
                self.cli_log_data["qth"] = user_input
                self.cli_log_step = "rst"
                self.print_to_terminal("请输入 RST [默认 59]:")
                
        elif step == "rst":
            self.cli_log_data["rst"] = user_input.strip() or "59"
            self.cli_log_step = "rig"
            self.show_options_for_input("Rig", "选择/输入设备 (Rig)")
            
        elif step == "rig":
            if not user_input.strip():  # 空输入处理
                self.print_to_terminal("设备不能为空，请重新输入:")
                return
                
            result = self.smart_match_input(user_input, "Rig")
            if result == "MULTIPLE_MATCHES":
                self.cli_log_step = "rig_select"
            elif result:
                self.cli_log_data["rig"] = result
                self.cli_log_step = "power"
                self.show_options_for_input("Power", "选择/输入功率 (Power)")
                
        elif step == "rig_select":
            self._handle_select("rig", user_input, "power", "Power")
                
        elif step == "power":
            # power允许空输入，使用默认值5W
            if not user_input.strip():
                self.cli_log_data["power"] = "5W"
                self.cli_log_step = "antenna"
                self.show_options_for_input("Antenna", "选择/输入天馈 (Antenna)")
                return
                
            result = self.smart_match_input(user_input, "Power")
            if result == "MULTIPLE_MATCHES":
                self.cli_log_step = "power_select"
            elif result:
                self.cli_log_data["power"] = result
                self.cli_log_step = "antenna"
                self.show_options_for_input("Antenna", "选择/输入天馈 (Antenna)")
            else:
                self.cli_log_data["power"] = user_input or "5W"
                self.cli_log_step = "antenna"
                self.show_options_for_input("Antenna", "选择/输入天馈 (Antenna)")
                
        elif step == "power_select":
            self._handle_select("power", user_input, "antenna", "Antenna", default="5W")
                
        elif step == "antenna":
            if not user_input.strip():  # 空输入处理
                self.print_to_terminal("天馈不能为空，请重新输入:")
                return
                
            result = self.smart_match_input(user_input, "Antenna")
            if result == "MULTIPLE_MATCHES":
                self.cli_log_step = "antenna_select"
            elif result:
                self.cli_log_data["antenna"] = result
                self.cli_log_step = "message"
                self.print_to_terminal("请输入讨论话题及留言 [默认 73]:")
                
        elif step == "antenna_select":
            self._handle_select("antenna", user_input, "message", "")
                
        elif step == "message":
            self.cli_log_data["message"] = user_input.strip() or "73"
            self.save_cli_log_record()

    def _handle_select(self, field_name, user_input, next_step, next_config_key, default=""):
        """处理多选情况的通用方法"""
        if user_input.isdigit():
            idx = int(user_input) - 1
            if 0 <= idx < len(self.current_matches):
                self.cli_log_data[field_name] = self.current_matches[idx]
                self.cli_log_step = next_step
                if next_step == "message":
                    self.print_to_terminal("请输入讨论话题及留言 [默认 73]:")
                else:
                    self.show_options_for_input(next_config_key, f"选择/输入{next_config_key} ({next_config_key})")
            else:
                self.print_to_terminal("序号无效，请重新选择:")
        else:
            self.cli_log_data[field_name] = user_input or default
            self.cli_log_step = next_step
            if next_step == "message":
                self.print_to_terminal("请输入讨论话题及留言 [默认 73]:")
            else:
                self.show_options_for_input(next_config_key, f"选择/输入{next_config_key} ({next_config_key})")

    def save_cli_log_record(self):
        """保存命令行录入的记录"""
        data = self.cli_log_data
        
        # 学习新词汇
        self.learn_new_value("QTH", data.get("qth", ""))
        self.learn_new_value("Rig", data.get("rig", ""))
        self.learn_new_value("Power", data.get("power", ""))
        self.learn_new_value("Antenna", data.get("antenna", ""))
        
        # 保存到Excel
        next_seq = self.ws.max_row
        current_time = datetime.datetime.now().strftime("%H:%M")
        
        self.ws.append([
            next_seq, 
            current_time, 
            data.get("callsign", ""), 
            data.get("qth", ""), 
            data.get("rst", "59"), 
            data.get("rig", ""), 
            data.get("power", "5W"), 
            data.get("antenna", ""), 
            data.get("message", "73")
        ])
        self.wb.save(EXCEL_FILE)
        
        # 同步导出CSV
        export_to_csv(self.ws)
        
        # 更新日志表格视图
        if self.log_tree is not None:
            self.log_tree.insert("", "end", values=(
                next_seq, current_time, data.get("callsign", ""), 
                data.get("qth", ""), data.get("rst", "59"), 
                data.get("rig", ""), data.get("power", "5W"), 
                data.get("antenna", ""), data.get("message", "73")
            ))
            children = self.log_tree.get_children()
            if children:
                self.log_tree.see(children[-1])
        
        # 显示确认信息
        self.print_to_terminal("-" * 35)
        self.print_to_terminal(f"✅ 已记录：{data.get('callsign', '')} | {data.get('qth', '')} | {data.get('rig', '')} | {data.get('power', '5W')} | {data.get('antenna', '')} | {data.get('rst', '59')} | {data.get('message', '73')}")
        self.print_to_terminal("-" * 35)
        
        # 重置状态
        self.cli_log_mode = False
        self.cli_log_data = {}
        self.cli_log_step = ""
        
        # 更新序号显示
        self.refresh_header()
        
        self.print_to_terminal("录入完成！输入 'log' 可继续录入下一条记录")


def main():
    root = tk.Tk()
    VibeLoggerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
